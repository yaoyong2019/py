from openpyxl import load_workbook
wb = load_workbook('test.xlsx')
print(wb.sheetnames)
sheet = wb.active
print(sheet.title)
A1 = sheet['A1']
print(A1.value)
B3_ = sheet.cell(row=1,column=2)
print(B3_.value)
""" print(sheet.max_column,sheet.max_row)
for column in sheet.columns:
    print(column) """
""" for i in range(1,4):
 for j in range(1,3):
  print(sheet.cell(row=i, column=j))
  print(i,j) """
""" print(sheet['A1:B3'])
for row_cell in sheet['A1:B3']:
    print(row_cell)
    for cell in row_cell:
        print (cell)
        print(cell.value) """
from openpyxl.utils import get_column_letter, column_index_from_string
# 根据列的数字返回字母
print(get_column_letter(2)) # B
# 根据字母返回列的数字
print(column_index_from_string('D')) # 4