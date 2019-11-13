import openpyxl
from openpyxl import Workbook
wb = openpyxl.load_workbook(r'f:\py\example1.xlsx')
sheet = wb.active
sheet.title = 'sheet1'
wb.create_sheet('data2',index=3)
wb.create_sheet('pdata2',index =4)
del wb['sheet1']
sheet = wb['data2']
sheet['A1'] = 'good'
sheet['A2'] = 123
c1 = sheet['A1']
c2 = sheet['A2']
row = [1 ,2, 3, 4, 5]
sheet.append(row)
# 添加多行
rows = [
 ['Number', 'data1', 'data2'],
 [2, 40, 30],
 [3, 40, 25],
 [4, 50, 30],
 [5.37, 30.40, 10.89],
 [60.0, 25.5, 5.0],
 ['7', '50', '10'],
 ['Ture','false','false'],
]
for row in rows:
    sheet.append(row)

from openpyxl.styles import Font, colors, Alignment
bold_itatic_24_font = Font(name='等线', size=24, italic=True, color=colors.RED, bold=True)
sheet['A1'].font = bold_itatic_24_font
sheet['A2'].alignment = Alignment(horizontal='center', vertical='center')
# 第2行行高
sheet.row_dimensions[2].height = 40
# C列列宽
sheet.column_dimensions['C'].width = 30
# 合并单元格， 往左上角写入数据即可
sheet.merge_cells('B1:G1') # 合并一行中的几个单元格
sheet.merge_cells('A4:C6') # 合并一个矩形区域中的单元格
sheet.unmerge_cells('B1:G1')
sheet.unmerge_cells('A4:C6')
wb.save(r'f:\py\example1.xlsx')